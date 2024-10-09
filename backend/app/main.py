from fastapi import FastAPI, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from .core.settings import settings
from .services.island_services import analyze_water_flow

import openpyxl
from io import BytesIO

app = FastAPI()

# add cors
app.add_middleware(CORSMiddleware,
    allow_origins=settings.cors_origins,  # Allow these origins
    allow_credentials=True,
    allow_methods=["*"],  # Allow all methods
    allow_headers=["*"]
    )

@app.get('/health')
def api_healt():
    return {
        "status": True,
        "message": "API health is good"
    }

@app.post('/upload')
async def upload_file(file: UploadFile = File(...)):
    file_contents = await file.read()
    excel_file = BytesIO(file_contents)
    workbook = openpyxl.load_workbook(excel_file, data_only = True)
    results = []

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        grid = []
        for row in worksheet.iter_rows(values_only=True):
            grid.append(list(row))
            
        
        # convert the grid into integers of 2D array
        grid = [[int(cell) for cell in row] for row in grid]
        print(grid)
        result = analyze_water_flow(grid)
        results.append({
            'sheet_name':sheet,
            'count':result
        })

    return results

